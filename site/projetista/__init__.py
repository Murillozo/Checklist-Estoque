# projetista/__init__.py
from flask import Blueprint, render_template, request, redirect, url_for, flash
from models import db, Solicitacao, Item, User, AuthorizedIP, EstoqueSolicitacao, EstoqueItem
from flask_login import login_required, current_user
import json
import io
from flask import send_file
import openpyxl
import pytz
from collections import Counter
from flask import jsonify
import os
from datetime import datetime
from werkzeug.utils import secure_filename
import urllib.parse
# PDF generation uses the fpdf2 package (pip install fpdf2) for Unicode support.
# Ensure that the TrueType font "DejaVuSans.ttf" is placed alongside this file.
from fpdf import FPDF
import re
LOGO_PATH = os.path.join(os.path.dirname(__file__), 'static', 'evomax_logo.png')


bp = Blueprint('projetista', __name__)

# Diretório onde os arquivos de checklist (JSON) são salvos.
# Permite sobrepor via variável de ambiente CHECKLIST_DIR.
CHECKLIST_DIR = os.environ.get(
    "CHECKLIST_DIR",
    os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'json_api')),
)

# Diretório base onde os projetos são armazenados no servidor
BASE_PRODUCAO = r"F:\03 - ENGENHARIA\03 - PRODUCAO"

# Diretório onde as fotos são salvas (pode ser sobrescrito via FOTOS_DIR)
# Por padrão aponta para a estrutura de projetos utilizada no servidor
FOTOS_DIR = os.environ.get(
    "FOTOS_DIR",
    r"F:\\03 - ENGENHARIA\\02 - PROJETOS"
)

# Garante que o diretório base exista para evitar erros de leitura
os.makedirs(FOTOS_DIR, exist_ok=True)

# Subpastas que devem ser criadas para cada obra
SUBPASTAS_OBRA = [
    'AS BUILT',
    'CHECKLIST',
    'FOTOS',
    'IDENTIFICAÇÕES',
    'LAYOUT',
    'PROJETO ELETROMECÂNICO',
]


@bp.route('/')
@login_required
def index():
    consulta = Solicitacao.query.order_by(Solicitacao.data.desc()).all()

    # Mantém apenas a ocorrência mais recente de cada obra
    unicas = []
    vistas = set()
    for sol in consulta:
        if sol.obra not in vistas:
            unicas.append(sol)
            vistas.add(sol.obra)

    # prepara lista de pendências para exibição
    for sol in unicas:
        try:
            sol.pendencias_list = json.loads(sol.pendencias or "[]")
        except json.JSONDecodeError:
            sol.pendencias_list = []

    return render_template('index.html', solicitacoes=unicas)

@bp.route('/solicitacoes')
@login_required
def solicitacoes():
    consulta = Solicitacao.query.order_by(Solicitacao.data.asc()).all()
    tz = pytz.timezone('America/Sao_Paulo')

    for sol in consulta:
        # converte horário
        utc_dt = sol.data.replace(tzinfo=pytz.UTC)
        sol.local_time = utc_dt.astimezone(tz)

        # agrupa itens repetidos
        agrupados = {}
        for it in sol.itens:
            agrupados[it.referencia] = agrupados.get(it.referencia, 0) + it.quantidade
        sol.itens_agrupados = list(agrupados.items())

    return render_template('solicitacoes.html', solicitacoes=consulta)


@bp.route('/iniciar_projeto')
@login_required
def iniciar_projeto():
    """Exibe os projetos divididos por status."""
    consulta = Solicitacao.query.order_by(Solicitacao.data.desc()).all()

    for sol in consulta:
        try:
            sol.pendencias_list = json.loads(sol.pendencias or "[]")
        except json.JSONDecodeError:
            sol.pendencias_list = []

    analise = [s for s in consulta if s.status == 'analise']
    aprovado = [s for s in consulta if s.status == 'aprovado']
    compras = [s for s in consulta if s.status == 'compras']

    return render_template(
        'iniciar_projeto.html',
        analise=analise,
        aprovado=aprovado,
        compras=compras,
    )


@bp.route('/solicitacao/nova', methods=['GET', 'POST'])
@login_required
def nova_solicitacao():
    # tenta descobrir os anos e obras disponíveis no servidor
    try:
        anos = [d for d in os.listdir(BASE_PRODUCAO)
                if os.path.isdir(os.path.join(BASE_PRODUCAO, d))]
        anos.sort()
        obras_por_ano = {}
        for ano in anos:
            ano_dir = os.path.join(BASE_PRODUCAO, ano)
            obras = [d for d in os.listdir(ano_dir)
                     if os.path.isdir(os.path.join(ano_dir, d))]
            obras.sort()
            obras_por_ano[ano] = obras
    except OSError:
        # se o diretório não estiver acessível, usa o ano atual
        ano_atual = str(datetime.now().year)
        anos = [ano_atual]
        obras_por_ano = {ano_atual: []}

    if request.method == 'POST':
        obra = request.form['obra'].strip()
        ano = request.form.get('ano', '').strip()
        data_entrega_str = request.form['data_entrega']
        data_entrega = datetime.strptime(data_entrega_str, '%Y-%m-%d').date()

        sol = Solicitacao(obra=obra, data_entrega=data_entrega)
        db.session.add(sol)
        db.session.flush()

        # 1) Se enviou um arquivo .xlsx, use ele:
        file = request.files.get('xlsx_file')
        if file and file.filename.lower().endswith(('.xls', '.xlsx')):
            wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)
            ws = wb.active
            # espera cabeçalho em linha 1: Referência | Quantidade
            for row in ws.iter_rows(min_row=2, values_only=True):
                ref, qt = row[0], row[1]
                if not ref or not qt:
                    continue
                item = Item(
                    solicitacao_id=sol.id,
                    referencia=str(ref).strip(),
                    quantidade=int(qt),
                    status='Nao iniciada'
                )
                db.session.add(item)

        # 2) Caso não tenha enviado arquivo, tenta textareas (legacy)
        else:
            refs = request.form.get('referencias', '').strip().splitlines()
            qts = request.form.get('quantidades', '').strip().splitlines()
            for ref, qt in zip(refs, qts):
                ref, qt = ref.strip(), qt.strip()
                if not ref or not qt:
                    continue
                item = Item(
                    solicitacao_id=sol.id,
                    referencia=ref,
                    quantidade=int(qt),
                    status='Nao iniciada'
                )
                db.session.add(item)

        db.session.commit()

        # cria as pastas da obra no servidor, se possível
        if ano:
            try:
                obra_dir = os.path.join(BASE_PRODUCAO, ano, obra)
                os.makedirs(obra_dir, exist_ok=True)

                # subpastas da obra principal
                for nome in SUBPASTAS_OBRA:
                    os.makedirs(os.path.join(obra_dir, nome), exist_ok=True)
            except OSError:
                # ignora falhas de criação de diretório
                pass

        flash('Solicitação criada com sucesso!', 'success')
        return redirect(url_for('projetista.solicitacoes'))

    return render_template('nova_solicitacao.html', anos=anos, obras_por_ano=obras_por_ano)


@bp.route('/verificar_estoque', methods=['GET', 'POST'])
@login_required
def verificar_estoque():
    if request.method == 'POST':
        itens = []
        file = request.files.get('xlsx_file')
        if file and file.filename.lower().endswith(('.xls', '.xlsx')):
            wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                ref, qt = row[0], row[1]
                if not ref or not qt:
                    continue
                itens.append({'referencia': str(ref).strip(), 'quantidade': int(qt)})
        else:
            refs = request.form.get('referencias', '').strip().splitlines()
            qts = request.form.get('quantidades', '').strip().splitlines()
            for ref, qt in zip(refs, qts):
                ref, qt = ref.strip(), qt.strip()
                if not ref or not qt:
                    continue
                itens.append({'referencia': ref, 'quantidade': int(qt)})

        if itens:
            sol = EstoqueSolicitacao()
            for it in itens:
                sol.itens.append(
                    EstoqueItem(
                        referencia=it['referencia'],
                        quantidade=it['quantidade']
                    )
                )
            db.session.add(sol)
            db.session.commit()
            flash('Solicitação registrada', 'success')
        return redirect(url_for('projetista.verificar_estoque'))

    solicitacoes = EstoqueSolicitacao.query.order_by(EstoqueSolicitacao.id.desc()).all()
    return render_template('verificar_estoque.html', solicitacoes=solicitacoes)


@bp.post('/verificar_estoque/<int:sol_id>/delete')
@login_required
def deletar_estoque_solicitacao(sol_id: int):
    """Remove uma solicitação de verificação de estoque."""
    sol = EstoqueSolicitacao.query.get_or_404(sol_id)
    db.session.delete(sol)
    db.session.commit()
    flash('Solicitação removida', 'success')
    return redirect(url_for('projetista.verificar_estoque'))


@bp.route('/subpastas', methods=['GET', 'POST'])
@login_required
def criar_subpastas():
    try:
        anos = [d for d in os.listdir(BASE_PRODUCAO)
                if os.path.isdir(os.path.join(BASE_PRODUCAO, d))]
    except OSError:
        anos = [str(datetime.now().year)]

    if request.method == 'POST':
        obra = request.form['obra'].strip()
        ano = request.form.get('ano', '').strip()
        subpastas_raw = request.form.get('subpastas', '0').strip()
        try:
            qtd_subpastas = int(subpastas_raw)
        except ValueError:
            qtd_subpastas = 0

        if ano and obra and qtd_subpastas > 0:
            try:
                obra_dir = os.path.join(BASE_PRODUCAO, ano, obra)
                for i in range(1, qtd_subpastas + 1):
                    sub_dir = os.path.join(obra_dir, f"{obra}.{i}")
                    for nome in SUBPASTAS_OBRA:
                        os.makedirs(os.path.join(sub_dir, nome), exist_ok=True)
                flash('Subpastas criadas com sucesso!', 'success')
            except OSError:
                flash('Falha ao criar subpastas.', 'danger')
        else:
            flash('Dados inválidos.', 'warning')
        return redirect(url_for('projetista.criar_subpastas'))

    return render_template('subpastas.html', anos=anos)


@bp.route('/comparador', methods=['GET', 'POST'])
@login_required
def comparador():
    # lista de obras únicas
    obras = [row[0] for row in db.session.query(Solicitacao.obra).distinct().all()]
    # todas as solicitações ordenadas por data ascendente
    solicitacoes = Solicitacao.query.order_by(Solicitacao.data.asc()).all()

    selecionada = None
    base = None
    nova = None
    adicionados = []
    removidos = []
    alterados = []

    if request.method == 'POST':
        selecionada = request.form.get('obra')
        id_base_raw = request.form.get('base')
        id_nova_raw = request.form.get('nova')

        # se só escolheu obra, mostra o form de IDs
        if selecionada and (not id_base_raw or not id_nova_raw):
            return render_template(
                'comparador.html',
                obras=obras,
                solicitacoes=solicitacoes,
                selecionada=selecionada
            )

        # converte para int
        id_base = int(id_base_raw)
        id_nova = int(id_nova_raw)

        # carrega objetos para comparar data
        sol_base = Solicitacao.query.get(id_base)
        sol_nova = Solicitacao.query.get(id_nova)

        # se base for mais recente que nova, troca
        if sol_base.data > sol_nova.data:
            sol_base, sol_nova = sol_nova, sol_base
            id_base, id_nova = id_nova, id_base

        # agrupa somando quantidades
        cnt_base = Counter()
        for it in Item.query.filter_by(solicitacao_id=sol_base.id):
            cnt_base[it.referencia] += it.quantidade

        cnt_nova = Counter()
        for it in Item.query.filter_by(solicitacao_id=sol_nova.id):
            cnt_nova[it.referencia] += it.quantidade

        # compara
        for ref in set(cnt_base) | set(cnt_nova):
            q0 = cnt_base.get(ref, 0)
            q1 = cnt_nova.get(ref, 0)
            if q0 == 0 and q1 > 0:
                adicionados.append((ref, q1))
            elif q1 == 0 and q0 > 0:
                removidos.append((ref, q0))
            elif q0 != q1:
                alterados.append((ref, q0, q1))

        return render_template(
            'comparador.html',
            obras=obras,
            solicitacoes=solicitacoes,
            selecionada=selecionada,
            base=id_base,
            nova=id_nova,
            adicionados=adicionados,
            removidos=removidos,
            alterados=alterados
        )

    # GET inicial
    return render_template(
        'comparador.html',
        obras=obras,
        solicitacoes=solicitacoes
    )

@bp.route('/template-solicitacao.xlsx')
def export_template():
    # Cria a planilha
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Solicitação"

    # Cabeçalhos
    ws.append(["Referência", "Quantidade"])

    # Você pode, opcionalmente, já deixar umas linhas de exemplo
    # ws.append(["ABC123", "10"])
    # ws.append(["DEF456", "5"])

    # Escreve num buffer em memória
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    # Envia como download
    return send_file(
        stream,
        as_attachment=True,
        download_name="template_solicitacao.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@bp.route('/api/solicitacoes')
@login_required
def api_listar_solicitacoes():
    resultados = []
    for sol in Solicitacao.query.order_by(Solicitacao.id.desc()).all():
        itens = [
            {
                "id": it.id,
                "referencia": it.referencia,
                "quantidade": it.quantidade,
                "status": it.status,
                "previsao_entrega": it.previsao_entrega.isoformat() if it.previsao_entrega else None,
            }
            for it in sol.itens
        ]
        resultados.append({
            "id": sol.id,
            "obra": sol.obra,
            "data": sol.data.isoformat(),
            "data_entrega": sol.data_entrega.isoformat(),
            "itens": itens,
            "status": sol.status,
            "pendencias": sol.pendencias
        })
    return jsonify(resultados)


@bp.route('/api/solicitacoes/<int:id>/aprovar', methods=['POST'])
@login_required
def api_aprovar(id):
    sol = Solicitacao.query.get_or_404(id)
    # Evita sobrescrever pendências já existentes quando a rota
    # de aprovação é chamada inadvertidamente. Assim, a solicitação
    # permanece com o status atual até que as pendências sejam
    # tratadas pelo setor responsável.
    if sol.pendencias and sol.pendencias != '[]':
        return jsonify({'ok': True})

    sol.status = 'aprovado'
    sol.pendencias = None
    db.session.commit()
    return jsonify({'ok': True})


@bp.route('/api/solicitacoes/<int:id>/compras', methods=['POST'])
@login_required
def api_compras(id):
    sol = Solicitacao.query.get_or_404(id)
    dados = request.get_json() or {}
    pendencias = dados.get('pendencias', [])

    # Se nenhuma pendência for enviada mas já existirem pendências
    # registradas, mantemos o status atual para evitar que a
    # solicitação seja marcada como completa de forma inadvertida.
    if not pendencias and sol.pendencias and sol.pendencias != '[]':
        return jsonify({'ok': True})

    total = len(sol.itens)
    concluido = total - len(pendencias)
    porcentagem = concluido / total if total else 0

    if porcentagem >= 0.8:
        sol.status = 'aprovado'
    elif sol.status != 'aprovado':
        sol.status = 'compras'
    sol.pendencias = json.dumps(pendencias)
    db.session.commit()
    return jsonify({'ok': True})


@bp.route('/checklist')
@login_required
def checklist_list():
    """Renderiza a interface de visualização dos checklists."""
    return render_template('checklist.html')


# =========================
# PDF COMPACTO (AGRUPADO)
# =========================
@bp.route('/checklist/pdf/<path:filename>')
@login_required
def checklist_pdf(filename):
    """Gera um PDF compacto (agrupado por Código + Item) a partir do checklist JSON."""
    caminho = os.path.join(CHECKLIST_DIR, filename)
    if not os.path.isfile(caminho):
        flash('Arquivo não encontrado.', 'danger')
        return redirect(url_for('projetista.checklist_list'))

    with open(caminho, encoding='utf-8') as f:
        dados = json.load(f)

    import unicodedata

    def _norm(s: str) -> str:
        s = (s or "").strip()
        s = ''.join(c for c in unicodedata.normalize('NFD', s)
                    if unicodedata.category(c) != 'Mn')
        s = s.upper().replace('—', ' ').replace('–', ' ').replace('-', ' ')
        return ' '.join(s.split())

    # ---------- Helpers de parsing/agrupamento ----------
    def _natural_key_codigo(pergunta: str):
        # pega "1.10" de "1.10 - CANALETAS: ..." e transforma em [1,10]
        norm = _norm(pergunta)
        if norm == "TENSAO CIRCUITO DE FORCA":
            return [4, 1, 5]
        m = re.match(r"\s*([0-9]+(?:\.[0-9]+)*)\s*-\s*", pergunta or "")
        if not m:
            return [float('inf')]
        return [int(p) for p in m.group(1).split('.')]

    def _split_pergunta(pergunta: str):
        # "1.1 - INVÓLUCRO - CAIXA: Identificação do projeto"
        # -> ("1.1", "INVÓLUCRO - CAIXA", "Identificação do projeto")
        pergunta = (pergunta or "").strip()
        codigo, item, sub = "", "", ""
        left, sub = (pergunta.split(":", 1) + [""])[:2]
        left = left.strip()
        sub = sub.strip()
        m = re.match(r"\s*([0-9]+(?:\.[0-9]+)*)\s*-\s*(.*)", left)
        if m:
            codigo = m.group(1).strip()
            item = m.group(2).strip()
        else:
            item = left
        return codigo, item, sub

    def _coletar_itens(node, acumulador):
        """Coleta recursivamente todos os itens 'pergunta' + 'respostas' em qualquer nível."""
        if isinstance(node, dict):
            lista = node.get('itens')
            if isinstance(lista, list):
                for it in lista:
                    pergunta = (it.get('pergunta') or "").strip()
                    # compat: respostas podem vir como dict/list/valor
                    brutas = it.get('respostas') or it.get('resposta', {})
                    resp_dict = {}
                    if isinstance(brutas, dict):
                        for k, v in brutas.items():
                            if isinstance(v, list):
                                resp_dict[k] = [str(x) for x in v if x is not None]
                            elif v is None:
                                resp_dict[k] = []
                            else:
                                resp_dict[k] = [str(v)]
                    elif isinstance(brutas, list):
                        resp_dict['resposta'] = [str(x) for x in brutas if x is not None]
                    elif brutas is not None:
                        resp_dict['resposta'] = [str(brutas)]
                    acumulador.append({'pergunta': pergunta, 'respostas': resp_dict})
                    _coletar_itens(it, acumulador)

            for k, v in node.items():
                if k != 'itens':
                    _coletar_itens(v, acumulador)
        elif isinstance(node, list):
            for elem in node:
                _coletar_itens(elem, acumulador)

    def _agrupar_por_codigo_item(items):
        """Agrupa subitens pelo par código + item para evitar repetições."""
        grupos = {}
        for it in sorted(items, key=lambda d: _natural_key_codigo(d.get('pergunta', ''))):
            codigo, item, sub = _split_pergunta(it.get('pergunta', ''))
            key = (codigo, item)
            grupos.setdefault(key, []).append({
                "subitem": sub,
                "respostas": it.get("respostas", {})
            })
        linhas = []
        for (codigo, item), subitens in grupos.items():
            linhas.append({
                "codigo": codigo,
                "item": item,
                "subitens": subitens,
                "respostas": [s["respostas"] for s in subitens]
            })
        return linhas

    # ---------- Montagem dos dados ----------
    planos = []
    _coletar_itens(dados, planos)
    grupos = _agrupar_por_codigo_item(planos)

    responsaveis = sorted({k for g in grupos
                           for resp in g["respostas"]
                           for k in resp})
    if not responsaveis:
        responsaveis = ["Suprimento", "Produção"]

    def _coletar_montadores(node):
        nomes = set()
        if isinstance(node, dict):
            for k, v in node.items():
                if k.lower() == "montador" and isinstance(v, str):
                    nome = v.strip()
                    if nome:
                        nomes.add(nome)
                else:
                    nomes.update(_coletar_montadores(v))
        elif isinstance(node, list):
            for elem in node:
                nomes.update(_coletar_montadores(elem))
        return nomes

    montadores = sorted(_coletar_montadores(dados))

    respondentes = dados.get("respondentes", {})
    suprimento = respondentes.get("suprimento", "").strip()
    producao = respondentes.get("produção", "").strip()

    # ---------- PDF ----------
    class ChecklistPDF(FPDF):
        def __init__(self, obra='', ano='', suprimento='', producao='', montadores=None, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self.obra = obra
            self.ano = ano
            self.suprimento = suprimento
            self.producao = producao
            self.montadores = montadores or []
            # aumenta margem inferior para que a tabela não sobreponha o rodapé
            self.set_auto_page_break(auto=False, margin=20)

        def header(self):
            self.set_fill_color(25, 25, 112)
            self.rect(0, 0, self.w, 25, 'F')
            self.set_y(5)
            if os.path.exists(LOGO_PATH):
                self.image(LOGO_PATH, x=10, y=5, w=40)
            self.set_text_color(255, 255, 255)
            self.set_font(base_font, 'B', 16)
            self.cell(0, 8, 'Checklist', align='C')
            self.set_font(base_font, '', 10)
            self.ln(6)
            self.cell(0, 5, f"Obra: {self.obra}   Ano: {self.ano}   Suprimento: {self.suprimento}   Produção: {self.producao}", align='C')
            self.ln(5)
            if self.montadores:
                nomes = ", ".join(f"{i+1}) {n}" for i, n in enumerate(self.montadores))
                self.cell(0, 5, f"Montadores: {nomes}", align='C')
                self.ln(5)
            self.set_y(40)
            self.set_text_color(0, 0, 0)

        def footer(self):
            self.set_y(-15)
            self.set_font(base_font, 'I', 8)
            self.set_text_color(128)
            self.cell(0, 10, f'Página {self.page_no()}/{{nb}}', align='C')

    pdf = ChecklistPDF(
        obra=dados.get('obra', ''),
        ano=dados.get('ano', ''),
        suprimento=suprimento,
        producao=producao,
        montadores=montadores,
        format='A4',
        orientation='P',
        unit='mm'
    )

    # Fontes (Unicode)
    # Coloque "DejaVuSans.ttf" ao lado deste arquivo (projetista/__init__.py)
    base_font = 'Arial'
    try:
        ttf_path = os.path.join(os.path.dirname(__file__), 'DejaVuSans.ttf')
        pdf.add_font('DejaVu', '', ttf_path, uni=True)
        pdf.add_font('DejaVu', 'B', ttf_path, uni=True)
        pdf.add_font('DejaVu', 'I', ttf_path, uni=True)
        base_font = 'DejaVu'
    except Exception:
        # fallback (pode perder acentos/símbolos)
        base_font = 'Arial'

    pdf.alias_nb_pages()
    pdf.add_page()
    pdf.set_font(base_font, size=7)


    bullet_char = "•" if base_font == "DejaVu" else "-"
    box_char = "□" if base_font == "DejaVu" else "[]"
    dash_char = "—" if base_font == "DejaVu" else "-"

    # ---------- Layout / medidas ----------
    left_margin = pdf.l_margin  # padrão 10 mm
    right_margin = pdf.r_margin
    usable_w = pdf.w - left_margin - right_margin

    # largura combinada para código + item + subitem
    col_w_item = 135.0
    # cada responsável ~22–28 mm
    col_w_resp = max(22.0, min(28.0, (usable_w - col_w_item) / max(1, len(responsaveis))))
    total_w = col_w_item + col_w_resp * len(responsaveis)
    if total_w > usable_w:
        # comprime a coluna de item proporcionalmente
        excesso = total_w - usable_w
        col_w_item = max(80.0, col_w_item - excesso)

    line_h = 6.0
    cell_pad = 2.0
    header_fill_rgb = (235, 235, 235)
    zebra_rgb = (247, 247, 247)

    def _wrap_lines(txt: str, width_mm: float):
        """Quebra em linhas para caber no width_mm atual (estimativa via get_string_width)."""
        txt = txt or ""
        if not txt:
            return [""]
        words = txt.split()
        lines, cur = [], ""
        for w in words:
            test = (cur + " " + w).strip()
            if pdf.get_string_width(test) <= width_mm - 2 * cell_pad:
                cur = test
            else:
                if cur:
                    lines.append(cur)
                cur = w
        if cur:
            lines.append(cur)
        return lines or [""]

    def _row_height(item_text):
        lines = []
        for line in (item_text or "").split("\n"):
            lines.extend(_wrap_lines(line, col_w_item))
        max_lines = max(len(lines), 1)
        return max(line_h * max_lines, line_h)

    def _header_row():
        x = left_margin
        y = pdf.get_y()
        pdf.set_fill_color(*header_fill_rgb)
        pdf.set_font(base_font, 'B', 10)
        # fundo do cabeçalho
        pdf.rect(x, y, col_w_item, line_h, 'F')
        cur_x = x + col_w_item
        for _ in responsaveis:
            pdf.rect(cur_x, y, col_w_resp, line_h, 'F')
            cur_x += col_w_resp
        # textos
        pdf.set_xy(x + cell_pad, y + 1)
        pdf.cell(col_w_item - 2 * cell_pad, line_h - 2, 'Item', border=0)
        cur_x = x + col_w_item
        for r in responsaveis:
            pdf.set_xy(cur_x + cell_pad, y + 1)
            pdf.cell(col_w_resp - 2 * cell_pad, line_h - 2, r.title(), border=0, align='C')
            cur_x += col_w_resp
        pdf.ln(line_h)
        pdf.set_font(base_font, '', 9)

    def _maybe_page_break(row_h, need_header=True):
        bottom_y = pdf.h - pdf.b_margin
        if pdf.get_y() + row_h > bottom_y:
            pdf.add_page()
            if need_header:
                _header_row()

    def _section_row(title: str):
        nonlocal zebra
        h = _row_height(title)
        top_gap = line_h
        _maybe_page_break(top_gap + h + line_h, need_header=False)
        pdf.ln(top_gap)
        pdf.set_fill_color(*header_fill_rgb)
        total_w = col_w_item + col_w_resp * len(responsaveis)
        pdf.rect(left_margin, pdf.get_y(), total_w, h, 'F')
        pdf.set_xy(left_margin + cell_pad, pdf.get_y() + 1)
        pdf.set_font(base_font, 'B', 10)
        pdf.cell(total_w - 2 * cell_pad, line_h - 2, title, border=0)
        pdf.ln(h)
        pdf.set_font(base_font, '', 10)
        _header_row()
        zebra = False

    # ---------- Tabela ----------
    sections_to_insert = [
        ("1.1", "INVOLUCRO CAIXA",               "POSTO - 01: MATERIAIS"),
        ("2.1", "PORTA",                         "POSTO - 02: OFICINA"),
        ("3.1", "COMPONENTE",                    "POSTO - 03: PRÉ-MONTAGEM - 01"),
        ("4.1", "BARRAMENTO",                    "POSTO - 04: BARRAMENTO - Identificação"),
        ("4.2", "COMANDO X TERRA",               "TESTE - TENSÃO APLICADA"),

        ("5.1", "CABLAGEM QD SOBREPOR/EMBUTIR",  "POSTO - 05: CABLAGEM - 01"),

        ("6.1", "COMPONENTES FIXACAO DIRETA",    "POSTO - 06: PRÉ-MONTAGEM - 02"),
        ("6.3", "CABLAGEM AUTOPORTANTE",         "POSTO - 06: CABLAGEM - 02"),
        ("",    "MULTIMEDIDOR",                  "TESTE - CONFIGURAÇÃO DE DISPOSITIVOS"),
        ("",    "SINALIZADOR",                   "TESTE - FUNCIONAIS"),
        ("",    "TORQUE PARAFUSOS DOS COMPONENTE","IQM - Inspeção de Qualidade Mecânica"),
        ("",    "CONTINUIDADE PONTO A PONTO FORCA","IQE - Inspeção de Qualidade Elétrica"),
        ("",    "RESPONSAVEL",                    "TESTES - DADOS"),
        ("",    "COMUNICADO A TRANSPORTADORA",    "EXPEDIÇÃO 01"),
        ("",    "LIMPEZA",                         "EXPEDIÇÃO 02"),
    ]
    inserted = set()
    zebra = False
    for g in grupos:
        codigo = g["codigo"] or ""
        item = g["item"] or dash_char
        base_item = f"{codigo} - {item}" if codigo else item
        item_norm = _norm(item)

        for cod_alvo, substr_item, titulo in sections_to_insert:
            key = (cod_alvo, substr_item, titulo)
            if key in inserted:
                continue
            if (cod_alvo and codigo.strip() == cod_alvo and substr_item in item_norm) or \
               (not cod_alvo and substr_item in item_norm):
                if titulo == "POSTO - 03: PRÉ-MONTAGEM - 01":
                    while pdf.page_no() < 5:
                        pdf.add_page()
                elif titulo == "POSTO - 04: BARRAMENTO - Identificação":
                    while pdf.page_no() < 6:
                        pdf.add_page()
                _section_row(titulo)
                inserted.add(key)

        subitens = g["subitens"] or [{"subitem": "", "respostas": {}}]

        for idx, sub in enumerate(subitens):
            item_text = base_item if idx == 0 else ""
            if sub["subitem"]:
                prefix = ("\n\n" if item_text else "")
                item_text += f"{prefix}{bullet_char} {sub['subitem']}"
            elif not item_text:
                item_text = dash_char

            roles_vals = []
            for role in responsaveis:
                vals = [str(v).strip() for v in sub["respostas"].get(role, []) if str(v).strip()]
                roles_vals.append(", ".join(vals) if vals else box_char)

            h = _row_height(item_text)
            _maybe_page_break(h)

            if zebra:
                pdf.set_fill_color(*zebra_rgb)
                pdf.rect(left_margin, pdf.get_y(), col_w_item + col_w_resp * len(responsaveis), h, 'F')
            zebra = not zebra

            x0 = left_margin
            y0 = pdf.get_y()
            pdf.rect(x0, y0, col_w_item, h)
            cur_x = x0 + col_w_item
            for _ in responsaveis:
                pdf.rect(cur_x, y0, col_w_resp, h)
                cur_x += col_w_resp

            pdf.set_xy(x0 + cell_pad, y0 + 1)
            pdf.multi_cell(col_w_item - 2 * cell_pad, line_h, item_text, border=0)

            cur_x = x0 + col_w_item
            for val in roles_vals:
                pdf.set_xy(cur_x, y0)
                pdf.cell(col_w_resp, h, val, border=0, align='C')
                cur_x += col_w_resp

            pdf.set_xy(left_margin, y0 + h)

    # Saída segura (fPDF2 em Py3 retorna bytes; se vier str, encode latin-1)
    out = pdf.output(dest='S')
    if isinstance(out, str):
        out = out.encode('latin-1', 'ignore')

    return send_file(
        io.BytesIO(out),
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'checklist_{dados.get("obra","")}_{dados.get("ano","")}_compacto.pdf'
    )




@bp.route('/checklist/<path:filename>')
@login_required
def checklist_view(filename):
    caminho = os.path.join(CHECKLIST_DIR, filename)
    if not os.path.isfile(caminho):
        flash('Arquivo não encontrado.', 'danger')
        return redirect(url_for('projetista.checklist_list'))
    with open(caminho, encoding='utf-8') as f:
        dados = json.load(f)
    # verifica se existe revisão anterior para comparação
    obra = dados.get('obra', 'Desconhecida') or 'Desconhecida'
    safe_obra = "".join(c for c in obra if c.isalnum() or c in ('-','_')) or 'obra'
    todos = [n for n in os.listdir(CHECKLIST_DIR)
            if n.endswith('.json') and n.startswith(f"checklist_{safe_obra}_")]
    todos.sort()
    try:
        idx = todos.index(filename)
        prev_filename = todos[idx - 1] if idx > 0 else None
    except ValueError:
        prev_filename = None
    return render_template(
        'checklist_view.html', filename=filename, dados=dados, prev_filename=prev_filename
    )


@bp.route('/checklist/diff/<path:filename>')
@login_required
def checklist_diff(filename):
    """Exibe as diferenças entre o checklist selecionado e o anterior."""
    caminho = os.path.join(CHECKLIST_DIR, filename)
    if not os.path.isfile(caminho):
        flash('Arquivo não encontrado.', 'danger')
        return redirect(url_for('projetista.checklist_list'))

    with open(caminho, encoding='utf-8') as f:
        atual = json.load(f)

    obra = atual.get('obra', 'Desconhecida') or 'Desconhecida'
    safe_obra = "".join(c for c in obra if c.isalnum() or c in ('-','_')) or 'obra'

    # Localiza o checklist anterior para a mesma obra
    todos = [n for n in os.listdir(CHECKLIST_DIR)
            if n.endswith('.json') and n.startswith(f"checklist_{safe_obra}_")]
    todos.sort()
    try:
        idx = todos.index(filename)
    except ValueError:
        idx = -1

    if idx <= 0:
        flash('Não há checklist anterior para comparação.', 'warning')
        return redirect(url_for('projetista.checklist_view', filename=filename))

    anterior_nome = todos[idx - 1]
    caminho_ant = os.path.join(CHECKLIST_DIR, anterior_nome)
    with open(caminho_ant, encoding='utf-8') as f:
        anterior = json.load(f)

    antigos = {i['pergunta']: i.get('resposta', [])
            for i in anterior.get('itens', [])}
    novos = {i['pergunta']: i.get('resposta', [])
            for i in atual.get('itens', [])}

    diff = []
    perguntas = sorted(set(antigos) | set(novos))
    for pergunta in perguntas:
        resp_ant = antigos.get(pergunta, [])
        resp_novo = novos.get(pergunta, [])
        if resp_ant != resp_novo:
            diff.append({
                'pergunta': pergunta,
                'antigo': ', '.join(map(str, resp_ant)),
                'novo': ', '.join(map(str, resp_novo))
            })

    return render_template(
        'checklist_diff.html',
        filename=filename,
        anterior=anterior_nome,
        diff=diff,
        obra=obra,
    )


@bp.route('/solicitacao/<int:id>/delete', methods=['POST'])
@login_required
def delete_solicitacao(id):
    sol = Solicitacao.query.get_or_404(id)
    db.session.delete(sol)
    db.session.commit()
    flash('Solicitação removida.', 'success')
    return redirect(url_for('projetista.index'))


@bp.route('/config', methods=['GET', 'POST'])
@login_required
def config():
    """Página para configurar usuários e senhas."""
    if current_user.role != 'admin':
        return redirect(url_for('projetista.index'))

    admin_user = User.query.filter_by(role='admin').first()
    compras_user = User.query.filter_by(role='compras').first()
    ips = AuthorizedIP.query.all()

    if request.method == 'POST':
        # Gerenciar adicao ou remocao de IPs
        new_ip = request.form.get('new_ip')
        delete_ip = request.form.get('delete_ip')
        if new_ip is not None:
            ip_val = new_ip.strip()
            if ip_val and not AuthorizedIP.query.filter_by(ip_address=ip_val).first():
                db.session.add(AuthorizedIP(ip_address=ip_val))
                db.session.commit()
                flash('IP adicionado.', 'success')
            return redirect(url_for('projetista.config'))
        if delete_ip:
            ip_obj = AuthorizedIP.query.get(int(delete_ip))
            if ip_obj:
                db.session.delete(ip_obj)
                db.session.commit()
                flash('IP removido.', 'success')
            return redirect(url_for('projetista.config'))

        admin_username = request.form.get('admin_username', '').strip()
        admin_password = request.form.get('admin_password', '').strip()
        compras_username = request.form.get('compras_username', '').strip()
        compras_password = request.form.get('compras_password', '').strip()

        if admin_username:
            if not admin_user:
                admin_user = User(username=admin_username, role='admin')
                db.session.add(admin_user)
            else:
                admin_user.username = admin_username
            if admin_password:
                admin_user.set_password(admin_password)

        if compras_username:
            if not compras_user:
                compras_user = User(username=compras_username, role='compras')
                db.session.add(compras_user)
            else:
                compras_user.username = compras_username
            if compras_password:
                compras_user.set_password(compras_password)

        db.session.commit()
        flash('Credenciais atualizadas.', 'success')
        return redirect(url_for('projetista.config'))

    return render_template(
        'config.html',
        admin_user=admin_user,
        compras_user=compras_user,
        ips=ips
    )


@bp.route('/api/inspecoes')
def api_inspecoes():
    dados = []
    for sol in EstoqueSolicitacao.query.filter_by(verificado=False).order_by(EstoqueSolicitacao.id.desc()).all():
        dados.append({
            'id': sol.id,
            'itens': [
                {
                    'id': i.id,
                    'referencia': i.referencia,
                    'quantidade': i.quantidade,
                    'verificado': i.verificado,
                    'faltante': i.faltante,
                }
                for i in sol.itens
            ]
        })
    return jsonify(dados)


@bp.route('/api/inspecoes/<int:id>/resultado', methods=['POST'])
def api_inspecao_resultado(id):
    sol = EstoqueSolicitacao.query.get_or_404(id)
    dados = request.get_json() or {}
    for item_data in dados.get('itens', []):
        item = EstoqueItem.query.get(item_data.get('id'))
        if item and item.solicitacao_id == id:
            item.verificado = item_data.get('verificado', False)
            item.faltante = item_data.get('faltante', 0)
    sol.verificado = True
    db.session.commit()
    return jsonify({'ok': True})


def _safe_join(root: str, *paths: str) -> str:
    root = os.path.abspath(root)
    path = os.path.abspath(os.path.join(root, *paths))
    if not path.startswith(root + os.sep):
        raise ValueError('Caminho inválido')
    return path


def _build_asbuilt_tree(base: str) -> list:
    """Percorre ``base`` e lista pastas ``AS BUILT/FOTOS`` e suas subpastas."""
    extensoes = (".jpg", ".jpeg", ".png")
    dados = {}
    alvo = os.path.join("AS BUILT", "FOTOS").upper()

    for root, _dirs, files in os.walk(base):
        if alvo not in root.upper():
            continue

        rel = os.path.relpath(root, base)
        partes = rel.split(os.sep)
        if len(partes) < 3:
            continue

        ano = partes[0]
        obra_path = "/".join(partes[1:])

        arquivos = [f for f in sorted(files) if f.lower().endswith(extensoes)]

        ano_dict = dados.setdefault(ano, {})
        ano_dict[obra_path] = [{'name': f} for f in arquivos]

    arvore = []
    for ano in sorted(dados.keys()):
        obras = []
        for obra in sorted(dados[ano].keys()):
            obras.append({'name': obra, 'children': dados[ano][obra]})
        arvore.append({'name': ano, 'children': obras})

    return arvore


@bp.route('/api/fotos')
def api_listar_fotos():
    """Lista pastas de fotos, incluindo subpastas vazias."""
    return jsonify(_build_asbuilt_tree(FOTOS_DIR))


@bp.route('/api/fotos/raw/<path:filepath>')
def api_foto_raw(filepath: str):
    filepath = urllib.parse.unquote(filepath)
    try:
        file_path = _safe_join(FOTOS_DIR, *filepath.split('/'))
    except ValueError:
        return jsonify({'error': 'Caminho inválido'}), 400
    if not os.path.isfile(file_path):
        return jsonify({'error': 'Arquivo não encontrado'}), 404
    return send_file(file_path)


@bp.route('/api/fotos/upload', methods=['POST'])
def api_enviar_foto():
    ano = request.form.get('ano', '').strip()
    obra = request.form.get('obra', '').strip()
    arquivo = request.files.get('foto')
    if not ano or not obra or not arquivo:
        return jsonify({'error': 'Dados incompletos'}), 400
    filename = secure_filename(arquivo.filename)
    try:
        destino = _safe_join(FOTOS_DIR, ano, *obra.split('/'))
        os.makedirs(destino, exist_ok=True)
    except ValueError:
        return jsonify({'error': 'Caminho inválido'}), 400
    caminho = os.path.join(destino, filename)
    arquivo.save(caminho)
    return jsonify({'ok': True})    
