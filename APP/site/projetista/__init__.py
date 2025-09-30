# projetista/__init__.py
from flask import Blueprint, render_template, request, redirect, url_for, flash
from models import db, Solicitacao, Item, User, AuthorizedIP, EstoqueSolicitacao, EstoqueItem
from flask_login import login_required, current_user
import json
import io
from flask import send_file
import openpyxl
import pytz
from collections import Counter
from flask import jsonify
import os
from datetime import datetime
from werkzeug.utils import secure_filename
import urllib.parse
# PDF generation uses the fpdf2 package (pip install fpdf2) for Unicode support.
# Ensure that the TrueType font "DejaVuSans.ttf" is placed alongside this file.
from fpdf import FPDF
from fontTools.ttLib import TTLibError
import logging
import re
LOGO_PATH = os.path.join(os.path.dirname(__file__), 'static', 'evomax_logo.png')


CHECKLIST_STATUS_MARKERS = {
    "C",
    "NC",
    "NA",
    "ND",
    "N/D",
    "OK",
    "P",
    "PEND",
    "PENDENTE",
}


def format_status_with_names(values, status_markers=None, status_separator=" — ", joiner=", "):
    """Return a human readable string for checklist answers.

    When the value starts with a status marker (e.g. ``C``/``NC``) followed by
    responder names, the status is highlighted separately from the names.
    ``status_separator`` controls how the status and the rest of the names are
    joined, while ``joiner`` is used for any additional names.
    """

    markers = status_markers or CHECKLIST_STATUS_MARKERS
    cleaned = []
    for raw in values or []:
        text = str(raw).strip()
        if text:
            cleaned.append(text)

    if not cleaned:
        return ""

    status, *names = cleaned
    printable_names = [name for name in names if name]
    if printable_names and status.upper() in markers:
        return f"{status}{status_separator}{joiner.join(printable_names)}"

    return joiner.join(cleaned)


def _identity(txt: str) -> str:
    return txt


def _latin1(txt: str) -> str:
    return txt.encode('latin-1', 'replace').decode('latin-1')


def draw_ohm(pdf: FPDF, x: float, y: float, size: float = 12) -> None:
    """Desenha o símbolo Ω usando a fonte "Symbol".

    Deve ser utilizado apenas quando o PDF foi gerado no modo de
    contingência (sem DejaVu), garantindo que o símbolo esteja presente
    mesmo sem suporte Unicode completo.
    """
    prev_family, prev_style, prev_size = pdf.font_family, pdf.font_style, pdf.font_size_pt
    pdf.set_font('Symbol', size=size)
    pdf.text(x, y, 'W')  # 'W' representa Ω na codificação da fonte Symbol
    pdf.set_font(prev_family, prev_style, prev_size)


def _load_dejavu(pdf: FPDF):
    """Tenta carregar DejaVuSans.ttf de diversos caminhos predefinidos.

    A ordem de busca é: diretório atual, ``static/`` local, diretórios
    ``projetista`` na raiz do projeto e ``site/projetista`` na raiz. Por
    fim, se a variável de ambiente ``DEJAVU_TTF`` estiver definida, o
    caminho informado é testado. Em caso de sucesso, as variantes
    regular, negrito e itálico são registradas no FPDF com suporte
    Unicode. Caso contrário, mantém-se o fallback para Arial.
    """
    tested = []
    root_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
    candidates = [
        os.path.join(os.path.dirname(__file__), 'DejaVuSans.ttf'),
        os.path.join(os.path.dirname(__file__), 'static', 'DejaVuSans.ttf'),
        os.path.join(root_dir, 'projetista', 'DejaVuSans.ttf'),
        os.path.join(root_dir, 'site', 'projetista', 'DejaVuSans.ttf'),
    ]
    env_path = os.environ.get('DEJAVU_TTF')
    if env_path:
        candidates.append(env_path)
    for path in candidates:
        tested.append(path)
        if not os.path.exists(path):
            continue
        try:
            pdf.add_font('DejaVu', '', path, uni=True)
            pdf.add_font('DejaVu', 'B', path, uni=True)
            pdf.add_font('DejaVu', 'I', path, uni=True)
            logging.info(f'Fonte DejaVuSans carregada: {path}')
            return 'DejaVu', _identity
        except (OSError, TTLibError):
            continue
    logging.warning('Não foi possível carregar DejaVuSans.ttf; caminhos testados: %s', tested)
    return 'Arial', _latin1



bp = Blueprint('projetista', __name__)

# Diretório onde os arquivos de checklist (JSON) são salvos.
# Permite sobrepor via variável de ambiente CHECKLIST_DIR.
CHECKLIST_DIR = os.environ.get(
    "CHECKLIST_DIR",
    os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'json_api')),
)

# Diretório base onde os projetos são armazenados no servidor
BASE_PRODUCAO = r"F:\03 - ENGENHARIA\03 - PRODUCAO"

# Diretório onde as fotos são salvas (pode ser sobrescrito via FOTOS_DIR)
# Por padrão aponta para a estrutura de projetos utilizada no servidor
FOTOS_DIR = os.environ.get(
    "FOTOS_DIR",
    r"F:\\03 - ENGENHARIA\\02 - PROJETOS"
)

# Garante que o diretório base exista para evitar erros de leitura
os.makedirs(FOTOS_DIR, exist_ok=True)

# Subpastas que devem ser criadas para cada obra
SUBPASTAS_OBRA = [
    'AS BUILT',
    'CHECKLIST',
    'FOTOS',
    'IDENTIFICAÇÕES',
    'LAYOUT',
    'PROJETO ELETROMECÂNICO',
]


@bp.route('/')
@login_required
def index():
    consulta = Solicitacao.query.order_by(Solicitacao.data.desc()).all()

    # Mantém apenas a ocorrência mais recente de cada obra
    unicas = []
    vistas = set()
    for sol in consulta:
        if sol.obra not in vistas:
            unicas.append(sol)
            vistas.add(sol.obra)

    # prepara lista de pendências para exibição
    for sol in unicas:
        try:
            sol.pendencias_list = json.loads(sol.pendencias or "[]")
        except json.JSONDecodeError:
            sol.pendencias_list = []

    return render_template('index.html', solicitacoes=unicas)

@bp.route('/solicitacoes')
@login_required
def solicitacoes():
    consulta = Solicitacao.query.order_by(Solicitacao.data.asc()).all()
    tz = pytz.timezone('America/Sao_Paulo')

    for sol in consulta:
        # converte horário
        utc_dt = sol.data.replace(tzinfo=pytz.UTC)
        sol.local_time = utc_dt.astimezone(tz)

        # agrupa itens repetidos
        agrupados = {}
        for it in sol.itens:
            agrupados[it.referencia] = agrupados.get(it.referencia, 0) + it.quantidade
        sol.itens_agrupados = list(agrupados.items())

    return render_template('solicitacoes.html', solicitacoes=consulta)


@bp.route('/iniciar_projeto')
@login_required
def iniciar_projeto():
    """Exibe os projetos divididos por status."""
    consulta = Solicitacao.query.order_by(Solicitacao.data.desc()).all()

    for sol in consulta:
        try:
            sol.pendencias_list = json.loads(sol.pendencias or "[]")
        except json.JSONDecodeError:
            sol.pendencias_list = []

    analise = [s for s in consulta if s.status == 'analise']
    aprovado = [s for s in consulta if s.status == 'aprovado']
    compras = [s for s in consulta if s.status == 'compras']

    return render_template(
        'iniciar_projeto.html',
        analise=analise,
        aprovado=aprovado,
        compras=compras,
    )


@bp.route('/solicitacao/nova', methods=['GET', 'POST'])
@login_required
def nova_solicitacao():
    # tenta descobrir os anos e obras disponíveis no servidor
    try:
        anos = [d for d in os.listdir(BASE_PRODUCAO)
                if os.path.isdir(os.path.join(BASE_PRODUCAO, d))]
        anos.sort()
        obras_por_ano = {}
        for ano in anos:
            ano_dir = os.path.join(BASE_PRODUCAO, ano)
            obras = [d for d in os.listdir(ano_dir)
                     if os.path.isdir(os.path.join(ano_dir, d))]
            obras.sort()
            obras_por_ano[ano] = obras
    except OSError:
        # se o diretório não estiver acessível, usa o ano atual
        ano_atual = str(datetime.now().year)
        anos = [ano_atual]
        obras_por_ano = {ano_atual: []}

    if request.method == 'POST':
        obra = request.form['obra'].strip()
        ano = request.form.get('ano', '').strip()
        data_entrega_str = request.form['data_entrega']
        data_entrega = datetime.strptime(data_entrega_str, '%Y-%m-%d').date()

        sol = Solicitacao(obra=obra, data_entrega=data_entrega)
        db.session.add(sol)
        db.session.flush()

        # 1) Se enviou um arquivo .xlsx, use ele:
        file = request.files.get('xlsx_file')
        if file and file.filename.lower().endswith(('.xls', '.xlsx')):
            wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)
            ws = wb.active
            # espera cabeçalho em linha 1: Referência | Quantidade
            for row in ws.iter_rows(min_row=2, values_only=True):
                ref, qt = row[0], row[1]
                if not ref or not qt:
                    continue
                item = Item(
                    solicitacao_id=sol.id,
                    referencia=str(ref).strip(),
                    quantidade=int(qt),
                    status='Separado'
                )
                db.session.add(item)

        # 2) Caso não tenha enviado arquivo, tenta textareas (legacy)
        else:
            refs = request.form.get('referencias', '').strip().splitlines()
            qts = request.form.get('quantidades', '').strip().splitlines()
            for ref, qt in zip(refs, qts):
                ref, qt = ref.strip(), qt.strip()
                if not ref or not qt:
                    continue
                item = Item(
                    solicitacao_id=sol.id,
                    referencia=ref,
                    quantidade=int(qt),
                    status='Separado'
                )
                db.session.add(item)

        db.session.commit()

        # cria as pastas da obra no servidor, se possível
        if ano:
            try:
                obra_dir = os.path.join(BASE_PRODUCAO, ano, obra)
                os.makedirs(obra_dir, exist_ok=True)

                # subpastas da obra principal
                for nome in SUBPASTAS_OBRA:
                    os.makedirs(os.path.join(obra_dir, nome), exist_ok=True)
            except OSError:
                # ignora falhas de criação de diretório
                pass

        flash('Solicitação criada com sucesso!', 'success')
        return redirect(url_for('projetista.solicitacoes'))

    return render_template('nova_solicitacao.html', anos=anos, obras_por_ano=obras_por_ano)


@bp.route('/verificar_estoque', methods=['GET', 'POST'])
@login_required
def verificar_estoque():
    if request.method == 'POST':
        itens = []
        file = request.files.get('xlsx_file')
        if file and file.filename.lower().endswith(('.xls', '.xlsx')):
            wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                ref, qt = row[0], row[1]
                if not ref or not qt:
                    continue
                itens.append({'referencia': str(ref).strip(), 'quantidade': int(qt)})
        else:
            refs = request.form.get('referencias', '').strip().splitlines()
            qts = request.form.get('quantidades', '').strip().splitlines()
            for ref, qt in zip(refs, qts):
                ref, qt = ref.strip(), qt.strip()
                if not ref or not qt:
                    continue
                itens.append({'referencia': ref, 'quantidade': int(qt)})

        if itens:
            sol = EstoqueSolicitacao()
            for it in itens:
                sol.itens.append(
                    EstoqueItem(
                        referencia=it['referencia'],
                        quantidade=it['quantidade']
                    )
                )
            db.session.add(sol)
            db.session.commit()
            flash('Solicitação registrada', 'success')
        return redirect(url_for('projetista.verificar_estoque'))

    solicitacoes = EstoqueSolicitacao.query.order_by(EstoqueSolicitacao.id.desc()).all()
    return render_template('verificar_estoque.html', solicitacoes=solicitacoes)


@bp.post('/verificar_estoque/<int:sol_id>/delete')
@login_required
def deletar_estoque_solicitacao(sol_id: int):
    """Remove uma solicitação de verificação de estoque."""
    sol = EstoqueSolicitacao.query.get_or_404(sol_id)
    db.session.delete(sol)
    db.session.commit()
    flash('Solicitação removida', 'success')
    return redirect(url_for('projetista.verificar_estoque'))


@bp.route('/subpastas', methods=['GET', 'POST'])
@login_required
def criar_subpastas():
    try:
        anos = [d for d in os.listdir(BASE_PRODUCAO)
                if os.path.isdir(os.path.join(BASE_PRODUCAO, d))]
    except OSError:
        anos = [str(datetime.now().year)]

    if request.method == 'POST':
        obra = request.form['obra'].strip()
        ano = request.form.get('ano', '').strip()
        subpastas_raw = request.form.get('subpastas', '0').strip()
        try:
            qtd_subpastas = int(subpastas_raw)
        except ValueError:
            qtd_subpastas = 0

        if ano and obra and qtd_subpastas > 0:
            try:
                obra_dir = os.path.join(BASE_PRODUCAO, ano, obra)
                for i in range(1, qtd_subpastas + 1):
                    sub_dir = os.path.join(obra_dir, f"{obra}.{i}")
                    for nome in SUBPASTAS_OBRA:
                        os.makedirs(os.path.join(sub_dir, nome), exist_ok=True)
                flash('Subpastas criadas com sucesso!', 'success')
            except OSError:
                flash('Falha ao criar subpastas.', 'danger')
        else:
            flash('Dados inválidos.', 'warning')
        return redirect(url_for('projetista.criar_subpastas'))

    return render_template('subpastas.html', anos=anos)


@bp.route('/comparador', methods=['GET', 'POST'])
@login_required
def comparador():
    # lista de obras únicas
    obras = [row[0] for row in db.session.query(Solicitacao.obra).distinct().all()]
    # todas as solicitações ordenadas por data ascendente
    solicitacoes = Solicitacao.query.order_by(Solicitacao.data.asc()).all()

    selecionada = None
    base = None
    nova = None
    adicionados = []
    removidos = []
    alterados = []

    if request.method == 'POST':
        selecionada = request.form.get('obra')
        id_base_raw = request.form.get('base')
        id_nova_raw = request.form.get('nova')

        # se só escolheu obra, mostra o form de IDs
        if selecionada and (not id_base_raw or not id_nova_raw):
            return render_template(
                'comparador.html',
                obras=obras,
                solicitacoes=solicitacoes,
                selecionada=selecionada
            )

        # converte para int
        id_base = int(id_base_raw)
        id_nova = int(id_nova_raw)

        # carrega objetos para comparar data
        sol_base = Solicitacao.query.get(id_base)
        sol_nova = Solicitacao.query.get(id_nova)

        # se base for mais recente que nova, troca
        if sol_base.data > sol_nova.data:
            sol_base, sol_nova = sol_nova, sol_base
            id_base, id_nova = id_nova, id_base

        # agrupa somando quantidades
        cnt_base = Counter()
        for it in Item.query.filter_by(solicitacao_id=sol_base.id):
            cnt_base[it.referencia] += it.quantidade

        cnt_nova = Counter()
        for it in Item.query.filter_by(solicitacao_id=sol_nova.id):
            cnt_nova[it.referencia] += it.quantidade

        # compara
        for ref in set(cnt_base) | set(cnt_nova):
            q0 = cnt_base.get(ref, 0)
            q1 = cnt_nova.get(ref, 0)
            if q0 == 0 and q1 > 0:
                adicionados.append((ref, q1))
            elif q1 == 0 and q0 > 0:
                removidos.append((ref, q0))
            elif q0 != q1:
                alterados.append((ref, q0, q1))

        return render_template(
            'comparador.html',
            obras=obras,
            solicitacoes=solicitacoes,
            selecionada=selecionada,
            base=id_base,
            nova=id_nova,
            adicionados=adicionados,
            removidos=removidos,
            alterados=alterados
        )

    # GET inicial
    return render_template(
        'comparador.html',
        obras=obras,
        solicitacoes=solicitacoes
    )

@bp.route('/template-solicitacao.xlsx')
def export_template():
    # Cria a planilha
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Solicitação"

    # Cabeçalhos
    ws.append(["Referência", "Quantidade"])

    # Você pode, opcionalmente, já deixar umas linhas de exemplo
    # ws.append(["ABC123", "10"])
    # ws.append(["DEF456", "5"])

    # Escreve num buffer em memória
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    # Envia como download
    return send_file(
        stream,
        as_attachment=True,
        download_name="template_solicitacao.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@bp.route('/api/solicitacoes')
@login_required
def api_listar_solicitacoes():
    resultados = []
    for sol in Solicitacao.query.order_by(Solicitacao.id.desc()).all():
        itens = [
            {
                "id": it.id,
                "referencia": it.referencia,
                "quantidade": it.quantidade,
                "status": it.status,
                "previsao_entrega": it.previsao_entrega.isoformat() if it.previsao_entrega else None,
            }
            for it in sol.itens
        ]
        resultados.append({
            "id": sol.id,
            "obra": sol.obra,
            "data": sol.data.isoformat(),
            "data_entrega": sol.data_entrega.isoformat(),
            "itens": itens,
            "status": sol.status,
            "pendencias": sol.pendencias
        })
    return jsonify(resultados)


@bp.route('/api/solicitacoes/<int:id>/aprovar', methods=['POST'])
@login_required
def api_aprovar(id):
    sol = Solicitacao.query.get_or_404(id)
    # Evita sobrescrever pendências já existentes quando a rota
    # de aprovação é chamada inadvertidamente. Assim, a solicitação
    # permanece com o status atual até que as pendências sejam
    # tratadas pelo setor responsável.
    if sol.pendencias and sol.pendencias != '[]':
        return jsonify({'ok': True})

    sol.status = 'aprovado'
    sol.pendencias = None
    db.session.commit()
    return jsonify({'ok': True})


@bp.route('/api/solicitacoes/<int:id>/compras', methods=['POST'])
@login_required
def api_compras(id):
    sol = Solicitacao.query.get_or_404(id)
    dados = request.get_json() or {}
    pendencias = dados.get('pendencias', [])

    # Se nenhuma pendência for enviada mas já existirem pendências
    # registradas, mantemos o status atual para evitar que a
    # solicitação seja marcada como completa de forma inadvertida.
    if not pendencias and sol.pendencias and sol.pendencias != '[]':
        return jsonify({'ok': True})

    total = len(sol.itens)
    concluido = total - len(pendencias)
    porcentagem = concluido / total if total else 0

    if porcentagem >= 0.8:
        sol.status = 'aprovado'
    elif sol.status != 'aprovado':
        sol.status = 'compras'
    sol.pendencias = json.dumps(pendencias)
    db.session.commit()
    return jsonify({'ok': True})


@bp.route('/checklist')
@login_required
def checklist_list():
    """Renderiza a interface de visualização dos checklists."""
    return render_template('checklist.html')


# =========================
# PDF COMPACTO (AGRUPADO)
# =========================
@bp.route('/checklist/pdf/<path:filename>')
@login_required
def checklist_pdf(filename):
    """Gera um PDF compacto (agrupado por Código + Item) a partir do checklist JSON."""
    caminho = os.path.join(CHECKLIST_DIR, filename)
    if not os.path.isfile(caminho):
        flash('Arquivo não encontrado.', 'danger')
        return redirect(url_for('projetista.checklist_list'))

    with open(caminho, encoding='utf-8') as f:
        dados = json.load(f)

    def _encontrar_inspetor(node):
        if isinstance(node, dict):
            v = node.get("inspetor")
            if isinstance(v, str) and v.strip():
                return v.strip()
            for _, val in node.items():
                res = _encontrar_inspetor(val)
                if res:
                    return res
        elif isinstance(node, list):
            for elem in node:
                res = _encontrar_inspetor(elem)
                if res:
                    return res
        return ""

    import unicodedata

    def _norm(s: str) -> str:
        s = (s or "").strip()
        s = ''.join(c for c in unicodedata.normalize('NFD', s)
                    if unicodedata.category(c) != 'Mn')
        s = s.upper().replace('—', ' ').replace('–', ' ').replace('-', ' ')
        return ' '.join(s.split())

    # ---------- Helpers de parsing/agrupamento ----------
    def _natural_key_codigo(pergunta: str):
        # pega "1.10" de "1.10 - CANALETAS: ..." e transforma em [1,10]
        m = re.match(r"\s*([0-9]+(?:\.[0-9]+)*)\s*-\s*", pergunta or "")
        if not m:
            return [float('inf')]
        return [int(p) for p in m.group(1).split('.')]

    def _split_pergunta(pergunta: str):
        # "1.1 - INVÓLUCRO - CAIXA: Identificação do projeto"
        # -> ("1.1", "INVÓLUCRO - CAIXA", "Identificação do projeto")
        pergunta = (pergunta or "").strip()
        codigo, item, sub = "", "", ""
        left, sub = (pergunta.split(":", 1) + [""])[:2]
        left = left.strip()
        sub = sub.strip()
        m = re.match(r"\s*([0-9]+(?:\.[0-9]+)*)\s*-\s*(.*)", left)
        if m:
            codigo = m.group(1).strip()
            item = m.group(2).strip()
        else:
            item = left
        return codigo, item, sub

    def _coletar_itens(node, acumulador):
        """Coleta recursivamente todos os itens 'pergunta' + 'respostas' em qualquer nível."""
        if isinstance(node, dict):
            lista = node.get('itens')
            if isinstance(lista, list):
                for it in lista:
                    pergunta = (it.get('pergunta') or "").strip()
                    # compat: respostas podem vir como dict/list/valor
                    brutas = it.get('respostas') or it.get('resposta', {})
                    resp_dict = {}
                    if isinstance(brutas, dict):
                        for k, v in brutas.items():
                            if isinstance(v, list):
                                resp_dict[k] = [str(x) for x in v if x is not None]
                            elif v is None:
                                resp_dict[k] = []
                            else:
                                resp_dict[k] = [str(v)]
                    elif isinstance(brutas, list):
                        resp_dict['resposta'] = [str(x) for x in brutas if x is not None]
                    elif brutas is not None:
                        resp_dict['resposta'] = [str(brutas)]
                    acumulador.append({'pergunta': pergunta, 'respostas': resp_dict})
                    _coletar_itens(it, acumulador)

            for k, v in node.items():
                if k != 'itens':
                    _coletar_itens(v, acumulador)
        elif isinstance(node, list):
            for elem in node:
                _coletar_itens(elem, acumulador)

    def _agrupar_por_codigo_item(items):
        """Agrupa subitens pelo par código + item para evitar repetições."""
        grupos = {}
        for it in sorted(items, key=lambda d: _natural_key_codigo(d.get('pergunta', ''))):
            codigo, item, sub = _split_pergunta(it.get('pergunta', ''))
            key = (codigo, item)
            grupos.setdefault(key, []).append({
                "subitem": sub,
                "respostas": it.get("respostas", {})
            })
        linhas = []
        for (codigo, item), subitens in grupos.items():
            linhas.append({
                "codigo": codigo,
                "item": item,
                "subitens": subitens,
                "respostas": [s["respostas"] for s in subitens]
            })
        return linhas

    # ---------- Montagem dos dados ----------
    planos = []
    _coletar_itens(dados, planos)
    grupos = _agrupar_por_codigo_item(planos)

    tensao_itens = [
        _norm("COMANDO X TERRA"),
        _norm("FORÇA - FASE A X BC TERRA"),
        _norm("FORÇA - FASE B X AC TERRA"),
        _norm("FORÇA - FASE C X AB TERRA"),
        _norm("FORÇA - FASE ABC X TERRA"),
    ]
    dados_itens = [
        _norm("RESPONSAVEL"),
        _norm("ALTITUDE EM RELAÇÃO AO NIVEL DO MAR"),
        _norm("GRAU DE POLUIÇÃO"),
        _norm("GRAU DE PROTEÇÃO (IP)"),
        _norm("INSTALAÇÃO"),
        _norm("APLICAÇÃO"),
        _norm("TEMPERATURA AMBIENTE"),
        _norm("HUMIDADE RELATIVA"),
        _norm("TENSÃO DE COMANDO"),
        _norm("TENSÃO CIRCUITO AUXILIAR"),
        _norm("TENSÃO CIRCUITO DE FORÇA"),
    ]

    idx_dados = idx_tensao_ini = idx_tensao_fim = None
    for i, g in enumerate(grupos):
        item_norm = _norm(g.get("item", ""))
        if item_norm in dados_itens:
            idx_dados = i
        if item_norm == tensao_itens[0] and idx_tensao_ini is None:
            idx_tensao_ini = i
        if item_norm == tensao_itens[-1]:
            idx_tensao_fim = i

    if None not in (idx_dados, idx_tensao_ini, idx_tensao_fim) and idx_tensao_ini <= idx_tensao_fim:
        bloco = grupos[idx_tensao_ini:idx_tensao_fim + 1]
        del grupos[idx_tensao_ini:idx_tensao_fim + 1]
        if idx_dados > idx_tensao_ini:
            idx_dados -= len(bloco)
        grupos[idx_dados + 1:idx_dados + 1] = bloco


    def _is_early_item(codigo: str) -> bool:
        parts = (codigo or "").split(".")
        if parts and parts[0] == "1" and len(parts) > 1:
            try:
                return 1 <= int(parts[1]) <= 14
            except ValueError:
                return False
        return False
    for g in grupos:
        if _is_early_item(g.get("codigo", "")):
            for resp in g.get("respostas", []):
                for k in list(resp.keys()):
                    v = resp[k]
                    norm_k = _norm(k)
                    if (
                        norm_k == "INSPETOR LOGISTICA MONTADOR PRODUCAO"
                        or v is None
                        or (isinstance(v, list) and not any(str(x).strip() for x in v))
                    ):
                        del resp[k]
            for sub in g.get("subitens", []):
                resp = sub.get("respostas", {})
                for k in list(resp.keys()):
                    v = resp[k]
                    norm_k = _norm(k)
                    if (
                        norm_k == "INSPETOR LOGISTICA MONTADOR PRODUCAO"
                        or v is None
                        or (isinstance(v, list) and not any(str(x).strip() for x in v))
                    ):
                        del resp[k]

    def _coletar_montadores(node):
        nomes = set()
        if isinstance(node, dict):
            for k, v in node.items():
                if k.lower() == "montador" and isinstance(v, str):
                    nome = v.strip()
                    if nome:
                        nomes.add(nome)
                else:
                    nomes.update(_coletar_montadores(v))
        elif isinstance(node, list):
            for elem in node:
                nomes.update(_coletar_montadores(elem))
        return nomes

    montadores = sorted(_coletar_montadores(dados))

    respondentes = dados.get("respondentes") or {}

    def _safe_responsavel(origem, *chaves):
        if not isinstance(origem, dict):
            return ""
        for chave in chaves:
            valor = origem.get(chave)
            if valor is None:
                continue
            texto = str(valor).strip()
            if texto:
                return texto
        return ""

    suprimento = _safe_responsavel(respondentes, "suprimento")
    producao = _safe_responsavel(respondentes, "produção", "producao")
    inspetor = _encontrar_inspetor(dados)

    def _safe_inspetor(section):
        if isinstance(section, dict):
            nome = section.get("inspetor")
            if isinstance(nome, str):
                return nome.strip()
        elif isinstance(section, str):
            return section.strip()
        return ""

    teste_insp = _safe_inspetor(dados.get("posto08_teste"))
    section_insp_names = {
        "POSTO - 02: OFICINA": _safe_inspetor(dados.get("posto02")),
        "POSTO - 03: PRÉ-MONTAGEM - 01": _safe_inspetor(
            dados.get("posto03_pre_montagem_01")
        ),
        "POSTO - 04: BARRAMENTO - Identificação": _safe_inspetor(
            dados.get("posto04_barramento")
        ),
        "POSTO - 05: CABLAGEM - 01": _safe_inspetor(
            dados.get("posto05_cablagem_01")
        ),
        "POSTO - 06: PRÉ-MONTAGEM - 02": _safe_inspetor(
            dados.get("posto06_pre_montagem_02")
        ),
        "POSTO - 06: CABLAGEM - 02": _safe_inspetor(
            dados.get("posto06_cablagem_02")
        ),
        "IQM - Inspeção de Qualidade Mecânica": _safe_inspetor(
            dados.get("posto08_iqm")
        ),
        "IQE - Inspeção de Qualidade Elétrica": _safe_inspetor(
            dados.get("posto08_iqe")
        ),
        "TESTE - CONFIGURAÇÃO DE DISPOSITIVOS": teste_insp,
        "TESTES - DADOS": teste_insp,
        "TESTE - FUNCIONAIS": teste_insp,
        "TESTE - TENSÃO APLICADA": teste_insp,
    }

    cidade_estado = request.args.get("cidade_estado", "").strip()
    if not cidade_estado:
        cidade = dados.get("cidade", "").strip()
        estado = dados.get("estado", "").strip()
        if cidade and estado:
            cidade_estado = f"{cidade}/{estado}"
        else:
            cidade_estado = cidade or estado

    projetista = request.args.get("projetista", "").strip()
    if not projetista:
        projetista = request.args.get("projesta", "").strip()
    if not projetista:
        projetista = str(dados.get("projetista") or dados.get("projesta") or "").strip()
    data_checklist = dados.get("data_checklist", datetime.now().strftime("%d/%m/%Y"))

    # ---------- PDF ----------
    safe_text = _identity

    class ChecklistPDF(FPDF):
        def __init__(self, obra='', ano='', suprimento='', producao='', montadores=None,
                     cidade_estado='', projetista='', projesta='', data_checklist='', inspetor='',
                     *args, **kwargs):
            super().__init__(*args, **kwargs)
            self.obra = obra
            self.ano = ano
            self.suprimento = suprimento
            self.producao = producao
            self.montadores = montadores or []
            self.cidade_estado = cidade_estado
            self.projetista = projetista or projesta
            self.projesta = self.projetista
            self.data_checklist = data_checklist
            self.inspetor = inspetor

        def header(self):
            # Faixa superior (azul) + logo + título
            self.set_fill_color(25, 25, 112)
            self.rect(0, 0, self.w, 25, 'F')
            if os.path.exists(LOGO_PATH):
                self.image(LOGO_PATH, x=10, y=5, w=40)

            self.set_text_color(255, 255, 255)
            self.set_font(base_font, 'B', 22)
            self.set_y(5.5)
            # Um pequeno espaçamento entre caracteres para dar um aspecto mais "futurista"
            set_char_spacing = getattr(self, 'set_char_spacing', None)
            if callable(set_char_spacing):
                set_char_spacing(0.75)
            self.cell(0, 11, 'CHECKLIST', align='C')
            if callable(set_char_spacing):
                set_char_spacing(0)

            # Só exibe o cartão com informações na primeira página
            if self.page_no() != 1:
                self.set_y(30)
                return

            # ---------- Cartão centralizado com os campos ----------
            info_items = [
                ("Cidade/Estado", self.cidade_estado or "-"),
                ("Obra",          self.obra or "-"),
                ("Ano",           str(self.ano or "-")),
                ("Data do Checklist", self.data_checklist or "-"),
                ("Projetista",    self.projetista or "-"),
                ("Inspetor",      self.inspetor or "-"),
                ("Suprimento",    self.suprimento or "-"),
                ("Produção",      self.producao or "-"),
            ]

            # Medidas do grid
            self.set_text_color(0, 0, 0)
            self.set_y(30)

            left_margin  = self.l_margin
            right_margin = self.r_margin
            usable_w     = self.w - left_margin - right_margin

            cols   = 4
            gap    = 3.0                 # espaçamento entre células
            grid_w = min(usable_w, 190)  # largura total do cartão
            cell_w = (grid_w - gap * (cols - 1)) / cols
            cell_h = 12.0                # altura de cada célula

            # Centraliza o cartão
            x0 = (self.w - grid_w) / 2.0
            y0 = self.get_y()

            # Moldura do cartão (leve)
            total_rows = (len(info_items) + cols - 1) // cols
            card_h = total_rows * cell_h + (total_rows - 1) * gap
            self.set_draw_color(210, 210, 210)
            self.set_fill_color(250, 250, 250)
            self.rect(x0 - 2, y0 - 2, grid_w + 4, card_h + 4, 'D')

            # Células (label pequeno + valor bold centralizado)
            for idx, (label, value) in enumerate(info_items):
                row = idx // cols
                col = idx % cols
                x = x0 + col * (cell_w + gap)
                y = y0 + row * (cell_h + gap)

                # Fundo da célula
                self.set_draw_color(230, 230, 230)
                self.set_fill_color(245, 245, 245)
                self.rect(x, y, cell_w, cell_h, 'DF')

                # Label
                self.set_font(base_font, '', 7)
                self.set_text_color(90, 90, 90)
                self.set_xy(x, y + 1.2)
                self.cell(cell_w, 3.2, label, border=0, align='C')

                # Valor
                self.set_font(base_font, 'B', 10)
                self.set_text_color(0, 0, 0)
                self.set_xy(x + 1, y + 4.4)
                self.cell(cell_w - 2, cell_h - 5, str(value), border=0, align='C')

            # Respiro abaixo do cartão antes da tabela
            self.set_y(y0 + card_h + 7)

        def footer(self):
            self.set_y(-15)
            self.set_font(base_font, 'I', 8)
            self.set_text_color(128)
            self.cell(0, 10, f'Página {self.page_no()}/{{nb}}', align='C')

        def cell(self, *args, **kwargs):
            if args:
                args = list(args)
                if len(args) >= 3 and isinstance(args[2], str):
                    args[2] = safe_text(args[2])
                args = tuple(args)
            if 'txt' in kwargs and isinstance(kwargs['txt'], str):
                kwargs['txt'] = safe_text(kwargs['txt'])
            return super().cell(*args, **kwargs)

        def multi_cell(self, *args, **kwargs):
            if args:
                args = list(args)
                if len(args) >= 3 and isinstance(args[2], str):
                    args[2] = safe_text(args[2])
                args = tuple(args)
            if 'txt' in kwargs and isinstance(kwargs['txt'], str):
                kwargs['txt'] = safe_text(kwargs['txt'])
            return super().multi_cell(*args, **kwargs)


    pdf = ChecklistPDF(
        obra=dados.get('obra', ''),
        ano=dados.get('ano', ''),
        suprimento=suprimento,
        producao=producao,
        montadores=montadores,
        cidade_estado=cidade_estado,
        projetista=projetista,
        data_checklist=data_checklist,
        inspetor=inspetor,
        format='A4',
        orientation='P',
        unit='mm'
    )

    pdf.set_margins(left=6, top=35, right=6)
    pdf.set_auto_page_break(auto=False, margin=20)

    base_font, safe_text = _load_dejavu(pdf)

    pdf.alias_nb_pages()
    pdf.add_page()
    pdf.set_font(base_font, size=7)


    bullet_char = "•" if base_font == "DejaVu" else "-"
    box_char = "□" if base_font == "DejaVu" else "[]"
    dash_char = "—" if base_font == "DejaVu" else "-"

    STATUS_MARKERS = CHECKLIST_STATUS_MARKERS
    NAME_ROLES = {"montador", "suprimento", "produção", "producao", "inspetor"}
    ROLE_LABELS = {
        "suprimento": "Suprimento",
        "produção": "Produção",
        "producao": "Produção",
        "montador": "Montador",
        "inspetor": "Inspetor",
        "logistica": "Logística",
        "logística": "Logística",
        "resposta": "Resposta",
    }

    def _is_potential_name(valor: str) -> bool:
        if not valor:
            return False
        texto = valor.strip()
        if not texto:
            return False
        if texto.upper() in STATUS_MARKERS:
            return False
        if any(ch.isdigit() for ch in texto):
            return False
        letras = sum(1 for ch in texto if ch.isalpha())
        return letras >= 3

    def _apply_last_name_rule(valores):
        name_indices = [idx for idx, val in enumerate(valores) if _is_potential_name(val)]
        if len(name_indices) < 2:
            return valores
        last_idx = name_indices[-1]
        filtrados = []
        for idx, val in enumerate(valores):
            if idx == last_idx:
                filtrados.append(val)
            elif idx in name_indices[:-1]:
                continue
            else:
                filtrados.append(val)
        return filtrados

    def _prepare_role_values(role, raw_values):
        if isinstance(raw_values, list):
            valores = [str(v).strip() for v in raw_values if str(v).strip()]
        elif raw_values is None:
            valores = []
        else:
            texto = str(raw_values).strip()
            valores = [texto] if texto else []
        role_lower = (role or "").lower()
        if role_lower in NAME_ROLES:
            valores = _apply_last_name_rule(valores)
        return valores

    # ---------- Layout / medidas ----------
    left_margin = pdf.l_margin
    right_margin = pdf.r_margin
    usable_w = pdf.w - left_margin - right_margin

    col_w_item = col_w_resp = total_w = 0.0

    line_h = 6.0
    cell_pad = 2.0
    header_fill_rgb = (235, 235, 235)
    zebra_rgb = (247, 247, 247)

    def _calc_widths(responsaveis_atual):
        count = len(responsaveis_atual)
        if not count:
            return usable_w, 0.0, usable_w
        col_w_item = 135.0
        col_w_resp = (usable_w - col_w_item) / count
        if count > 1 and col_w_resp > 28.0:
            col_w_resp = 28.0
            col_w_item = usable_w - col_w_resp * count
        elif col_w_resp < 22.0:
            col_w_resp = 22.0
            col_w_item = usable_w - col_w_resp * count
            if col_w_item < 80.0:
                col_w_item = 80.0
                col_w_resp = (usable_w - col_w_item) / count
        total_w = usable_w
        return col_w_item, col_w_resp, total_w

    def _wrap_lines(txt: str, width_mm: float):
        """Quebra em linhas para caber no width_mm atual (estimativa via get_string_width)."""
        txt = txt or ""
        if not txt:
            return [""]
        words = txt.split()
        lines, cur = [], ""
        for w in words:
            test = (cur + " " + w).strip()
            if pdf.get_string_width(test) <= width_mm - 2 * cell_pad:
                cur = test
            else:
                if cur:
                    lines.append(cur)
                cur = w
        if cur:
            lines.append(cur)
        return lines or [""]

    def _count_lines(text: str, width_mm: float) -> int:
        lines = []
        for line in (text or "").split("\n"):
            lines.extend(_wrap_lines(line, width_mm))
        return max(len(lines), 1)

    def _row_height(item_text):
        return max(line_h * _count_lines(item_text, col_w_item), line_h)

    current_roles = []
    current_section = ""

    def _header_row(responsaveis_atual):
        _maybe_page_break(line_h, need_header=False)
        x = left_margin
        y = pdf.get_y()
        pdf.set_fill_color(*header_fill_rgb)
        pdf.set_font(base_font, 'B', 10)
        pdf.rect(x, y, col_w_item, line_h, 'F')
        cur_x = x + col_w_item
        for _ in responsaveis_atual:
            pdf.rect(cur_x, y, col_w_resp, line_h, 'F')
            cur_x += col_w_resp
        pdf.set_xy(x + cell_pad, y + 1)
        pdf.cell(col_w_item - 2 * cell_pad, line_h - 2, 'Item', border=0)
        cur_x = x + col_w_item
        for r in responsaveis_atual:
            role_lower = (r or "").lower()
            header_txt = ROLE_LABELS.get(role_lower, r.title())
            if role_lower == "inspetor":
                insp_name = section_insp_names.get(current_section, "").strip()
                if insp_name:
                    header_txt = f"Inspetor: {insp_name}"
            pdf.set_xy(cur_x + cell_pad, y + 1)
            pdf.cell(col_w_resp - 2 * cell_pad, line_h - 2, header_txt, border=0, align='C')
            cur_x += col_w_resp
        pdf.ln(line_h)
        pdf.set_font(base_font, '', 9)

    def _maybe_page_break(row_h, need_header=True):
        bottom_y = pdf.h - pdf.b_margin
        if pdf.get_y() + row_h > bottom_y:
            pdf.add_page()
            if need_header and current_roles:
                _header_row(current_roles)

    def _section_row(title: str, responsaveis_atual):
        nonlocal zebra, current_section
        h = _row_height(title)
        extra = line_h  # espaçamento adicional antes do título
        _maybe_page_break(extra + h + line_h, need_header=False)
        pdf.ln(extra)
        pdf.set_fill_color(*header_fill_rgb)
        y0 = pdf.get_y()
        pdf.rect(left_margin, y0, total_w, h, 'F')
        pdf.set_xy(left_margin + cell_pad, y0 + 1)
        pdf.set_font(base_font, 'B', 10)
        pdf.cell(total_w - 2 * cell_pad, line_h - 2, title, border=0)
        pdf.ln(h)
        pdf.set_font(base_font, '', 7)
        zebra = False
        current_section = title

    def _roles_present_in_group(g):
        roles = set()
        for sub in g.get("subitens", []):
            for k, v in (sub.get("respostas") or {}).items():
                if isinstance(v, list):
                    if any(str(x).strip() for x in v):
                        roles.add(k)
                else:
                    if str(v).strip():
                        roles.add(k)
        return sorted(roles)

    # ---------- Tabela ----------
    sections_to_insert = [
        ("1.1", "INVOLUCRO CAIXA",               "POSTO - 01: MATERIAIS"),
        ("2.1", "PORTA",                         "POSTO - 02: OFICINA"),
        ("3.1", "COMPONENTE",                    "POSTO - 03: PRÉ-MONTAGEM - 01"),
        ("4.1", "BARRAMENTO",                    "POSTO - 04: BARRAMENTO - Identificação"),

        ("5.1", "CABLAGEM QD SOBREPOR/EMBUTIR",  "POSTO - 05: CABLAGEM - 01"),

        ("6.1", "COMPONENTES FIXACAO DIRETA",    "POSTO - 06: PRÉ-MONTAGEM - 02"),
        ("6.3", "CABLAGEM AUTOPORTANTE",         "POSTO - 06: CABLAGEM - 02"),
        ("",    "MULTIMEDIDOR",                  "TESTE - CONFIGURAÇÃO DE DISPOSITIVOS"),
        ("",    "SINALIZADOR",                   "TESTE - FUNCIONAIS"),
        ("",    "TORQUE PARAFUSOS DOS COMPONENTE","IQM - Inspeção de Qualidade Mecânica"),
        ("",    "CONTINUIDADE PONTO A PONTO FORCA","IQE - Inspeção de Qualidade Elétrica"),
        ("",    "RESPONSAVEL",                    "TESTES - DADOS"),
        ("4.2", "COMANDO X TERRA",               "TESTE - TENSÃO APLICADA"),
        ("",    "COMUNICADO A TRANSPORTADORA",    "EXPEDIÇÃO 01"),
        ("",    "LIMPEZA",                         "EXPEDIÇÃO 02"),
    ]
    inserted = set()
    zebra = False
    for g in grupos:
        codigo = g["codigo"] or ""
        item = g["item"] or dash_char
        base_item = f"{codigo} - {item}" if codigo else item
        item_norm = _norm(item)

        roles_this = _roles_present_in_group(g)
        if _is_early_item(codigo) and not roles_this:
            roles_this = ["suprimento"]
        header_needed = False
        if roles_this != current_roles:
            col_w_item, col_w_resp, total_w = _calc_widths(roles_this)
            current_roles = roles_this
            header_needed = True
            zebra = False

        for cod_alvo, substr_item, titulo in sections_to_insert:
            key = (cod_alvo, substr_item, titulo)
            if key in inserted:
                continue
            if (cod_alvo and codigo.strip() == cod_alvo and substr_item in item_norm) or \
               (not cod_alvo and substr_item in item_norm):
                if titulo == "POSTO - 03: PRÉ-MONTAGEM - 01":
                    while pdf.page_no() < 5:
                        pdf.add_page()
                elif titulo == "POSTO - 04: BARRAMENTO - Identificação":
                    while pdf.page_no() < 6:
                        pdf.add_page()
                _section_row(titulo, roles_this)
                header_needed = True
                inserted.add(key)

        if header_needed:
            _header_row(current_roles)

        subitens = g["subitens"] or [{"subitem": "", "respostas": {}}]

        for idx, sub in enumerate(subitens):
            item_text = base_item if idx == 0 else ""
            if sub["subitem"]:
                prefix = ("\n" if item_text else "")
                item_text += f"{prefix}{bullet_char} {sub['subitem']}"
            elif not item_text:
                item_text = dash_char

            roles_vals = []
            max_resp_lines = 0
            for role in current_roles:
                role_vals = (sub.get("respostas") or {}).get(role)
                vals = _prepare_role_values(role, role_vals)
                role_lower = (role or "").lower()
                if role_lower == "inspetor":
                    insp_name = section_insp_names.get(current_section, "").strip()
                    if insp_name:
                        normalized_insp = _norm(insp_name)
                        existing_names = [
                            _norm(v)
                            for v in vals
                            if v and v.upper() not in STATUS_MARKERS
                        ]
                        if normalized_insp and normalized_insp not in existing_names:
                            vals = list(vals)
                            if vals and vals[0].upper() in STATUS_MARKERS:
                                vals.append(insp_name)
                            elif vals:
                                vals.append(insp_name)
                            else:
                                vals = [insp_name]
                if role_lower in ("resposta", "inspetor") and len(vals) >= 5:
                    formatted = (
                        f"1. Tensão aplicada: {vals[1]} {vals[0]}\n"
                        f"2. Resultado: {vals[3]} {vals[2]}\n"
                        f"3. Situação: {vals[4]}"
                    )
                elif len(vals) >= 5:
                    formatted = "\n".join(f"{i+1}. {v}" for i, v in enumerate(vals))
                else:
                    formatted = format_status_with_names(vals, STATUS_MARKERS)
                if not formatted:
                    formatted = box_char
                roles_vals.append(formatted)
                max_resp_lines = max(max_resp_lines, _count_lines(formatted, col_w_resp))

            item_lines = _count_lines(item_text, col_w_item)
            h = line_h * max(item_lines, max_resp_lines)
            _maybe_page_break(h)

            if zebra:
                pdf.set_fill_color(*zebra_rgb)
                pdf.rect(left_margin, pdf.get_y(), total_w, h, 'F')
            zebra = not zebra

            x0 = left_margin
            y0 = pdf.get_y()
            pdf.rect(x0, y0, col_w_item, h)
            cur_x = x0 + col_w_item
            for _ in current_roles:
                pdf.rect(cur_x, y0, col_w_resp, h)
                cur_x += col_w_resp

            pdf.set_xy(x0 + cell_pad, y0 + 1)
            pdf.multi_cell(col_w_item - 2 * cell_pad, line_h, item_text, border=0)

            cur_x = x0 + col_w_item
            for val in roles_vals:
                pdf.set_xy(cur_x + cell_pad, y0 + 1)
                pdf.multi_cell(col_w_resp - 2 * cell_pad, line_h, val, border=0, align='C')
                cur_x += col_w_resp
                pdf.set_xy(cur_x, y0)

            pdf.set_xy(left_margin, y0 + h)

    # Saída segura (fPDF2 em Py3 retorna bytes; se vier str, encode latin-1)
    out = pdf.output(dest='S')
    if isinstance(out, str):
        out = out.encode('latin-1', 'ignore')

    return send_file(
        io.BytesIO(out),
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'checklist_{dados.get("obra","")}_{dados.get("ano","")}_compacto.pdf'
    )




@bp.route('/checklist/<path:filename>')
@login_required
def checklist_view(filename):
    caminho = os.path.join(CHECKLIST_DIR, filename)
    if not os.path.isfile(caminho):
        flash('Arquivo não encontrado.', 'danger')
        return redirect(url_for('projetista.checklist_list'))
    with open(caminho, encoding='utf-8') as f:
        dados = json.load(f)
    # Extrai apenas as respostas mais recentes e remove valores nulos
    itens_brutos = dados.get('itens', [])
    itens_processados = []
    for item in itens_brutos:
        respostas_dict = item.get('respostas') or {}
        respostas_coletadas = []
        for pessoa, respostas in respostas_dict.items():
            ultima = None
            if isinstance(respostas, list):
                for r in reversed(respostas):
                    if r is not None:
                        ultima = r
                        break
            else:
                if respostas is not None:
                    ultima = respostas
            if ultima is not None:
                respostas_coletadas.append(f"{pessoa}: {ultima}")
        if respostas_coletadas:
            itens_processados.append({
                'numero': item.get('numero'),
                'pergunta': item.get('pergunta'),
                'resposta': ', '.join(respostas_coletadas)
            })
    itens_processados.sort(key=lambda x: x.get('numero', 0), reverse=True)
    dados['itens'] = itens_processados
    # verifica se existe revisão anterior para comparação
    obra = dados.get('obra', 'Desconhecida') or 'Desconhecida'
    safe_obra = "".join(c for c in obra if c.isalnum() or c in ('-','_')) or 'obra'
    todos = [n for n in os.listdir(CHECKLIST_DIR)
            if n.endswith('.json') and n.startswith(f"checklist_{safe_obra}_")]
    todos.sort()
    try:
        idx = todos.index(filename)
        prev_filename = todos[idx - 1] if idx > 0 else None
    except ValueError:
        prev_filename = None
    return render_template(
        'checklist_view.html', filename=filename, dados=dados, prev_filename=prev_filename
    )


@bp.route('/checklist/diff/<path:filename>')
@login_required
def checklist_diff(filename):
    """Exibe as diferenças entre o checklist selecionado e o anterior."""
    caminho = os.path.join(CHECKLIST_DIR, filename)
    if not os.path.isfile(caminho):
        flash('Arquivo não encontrado.', 'danger')
        return redirect(url_for('projetista.checklist_list'))

    with open(caminho, encoding='utf-8') as f:
        atual = json.load(f)

    obra = atual.get('obra', 'Desconhecida') or 'Desconhecida'
    safe_obra = "".join(c for c in obra if c.isalnum() or c in ('-','_')) or 'obra'

    # Localiza o checklist anterior para a mesma obra
    todos = [n for n in os.listdir(CHECKLIST_DIR)
            if n.endswith('.json') and n.startswith(f"checklist_{safe_obra}_")]
    todos.sort()
    try:
        idx = todos.index(filename)
    except ValueError:
        idx = -1

    if idx <= 0:
        flash('Não há checklist anterior para comparação.', 'warning')
        return redirect(url_for('projetista.checklist_view', filename=filename))

    anterior_nome = todos[idx - 1]
    caminho_ant = os.path.join(CHECKLIST_DIR, anterior_nome)
    with open(caminho_ant, encoding='utf-8') as f:
        anterior = json.load(f)

    antigos = {i['pergunta']: i.get('resposta', [])
            for i in anterior.get('itens', [])}
    novos = {i['pergunta']: i.get('resposta', [])
            for i in atual.get('itens', [])}

    diff = []
    perguntas = sorted(set(antigos) | set(novos))
    for pergunta in perguntas:
        resp_ant = antigos.get(pergunta, [])
        resp_novo = novos.get(pergunta, [])
        if resp_ant != resp_novo:
            diff.append({
                'pergunta': pergunta,
                'antigo': ', '.join(map(str, resp_ant)),
                'novo': ', '.join(map(str, resp_novo))
            })

    return render_template(
        'checklist_diff.html',
        filename=filename,
        anterior=anterior_nome,
        diff=diff,
        obra=obra,
    )


@bp.route('/solicitacao/<int:id>/delete', methods=['POST'])
@login_required
def delete_solicitacao(id):
    sol = Solicitacao.query.get_or_404(id)
    db.session.delete(sol)
    db.session.commit()
    flash('Solicitação removida.', 'success')
    return redirect(url_for('projetista.index'))


@bp.route('/config', methods=['GET', 'POST'])
@login_required
def config():
    """Página para configurar usuários e senhas."""
    if current_user.role != 'admin':
        return redirect(url_for('projetista.index'))

    admin_user = User.query.filter_by(role='admin').first()
    compras_user = User.query.filter_by(role='compras').first()
    ips = AuthorizedIP.query.all()

    if request.method == 'POST':
        # Gerenciar adicao ou remocao de IPs
        new_ip = request.form.get('new_ip')
        delete_ip = request.form.get('delete_ip')
        if new_ip is not None:
            ip_val = new_ip.strip()
            if ip_val and not AuthorizedIP.query.filter_by(ip_address=ip_val).first():
                db.session.add(AuthorizedIP(ip_address=ip_val))
                db.session.commit()
                flash('IP adicionado.', 'success')
            return redirect(url_for('projetista.config'))
        if delete_ip:
            ip_obj = AuthorizedIP.query.get(int(delete_ip))
            if ip_obj:
                db.session.delete(ip_obj)
                db.session.commit()
                flash('IP removido.', 'success')
            return redirect(url_for('projetista.config'))

        admin_username = request.form.get('admin_username', '').strip()
        admin_password = request.form.get('admin_password', '').strip()
        compras_username = request.form.get('compras_username', '').strip()
        compras_password = request.form.get('compras_password', '').strip()

        if admin_username:
            if not admin_user:
                admin_user = User(username=admin_username, role='admin')
                db.session.add(admin_user)
            else:
                admin_user.username = admin_username
            if admin_password:
                admin_user.set_password(admin_password)

        if compras_username:
            if not compras_user:
                compras_user = User(username=compras_username, role='compras')
                db.session.add(compras_user)
            else:
                compras_user.username = compras_username
            if compras_password:
                compras_user.set_password(compras_password)

        db.session.commit()
        flash('Credenciais atualizadas.', 'success')
        return redirect(url_for('projetista.config'))

    return render_template(
        'config.html',
        admin_user=admin_user,
        compras_user=compras_user,
        ips=ips
    )


@bp.route('/api/inspecoes')
def api_inspecoes():
    dados = []
    for sol in EstoqueSolicitacao.query.filter_by(verificado=False).order_by(EstoqueSolicitacao.id.desc()).all():
        dados.append({
            'id': sol.id,
            'itens': [
                {
                    'id': i.id,
                    'referencia': i.referencia,
                    'quantidade': i.quantidade,
                    'verificado': i.verificado,
                    'faltante': i.faltante,
                }
                for i in sol.itens
            ]
        })
    return jsonify(dados)


@bp.route('/api/inspecoes/<int:id>/resultado', methods=['POST'])
def api_inspecao_resultado(id):
    sol = EstoqueSolicitacao.query.get_or_404(id)
    dados = request.get_json() or {}
    for item_data in dados.get('itens', []):
        item = EstoqueItem.query.get(item_data.get('id'))
        if item and item.solicitacao_id == id:
            item.verificado = item_data.get('verificado', False)
            item.faltante = item_data.get('faltante', 0)
    sol.verificado = True
    db.session.commit()
    return jsonify({'ok': True})


def _safe_join(root: str, *paths: str) -> str:
    root = os.path.abspath(root)
    path = os.path.abspath(os.path.join(root, *paths))
    if not path.startswith(root + os.sep):
        raise ValueError('Caminho inválido')
    return path


def _build_asbuilt_tree(base: str) -> list:
    """Percorre ``base`` e lista pastas ``AS BUILT/FOTOS`` e suas subpastas."""
    extensoes = (".jpg", ".jpeg", ".png")
    dados = {}
    alvo = os.path.join("AS BUILT", "FOTOS").upper()

    for root, _dirs, files in os.walk(base):
        if alvo not in root.upper():
            continue

        rel = os.path.relpath(root, base)
        partes = rel.split(os.sep)
        if len(partes) < 3:
            continue

        ano = partes[0]
        obra_path = "/".join(partes[1:])

        arquivos = [f for f in sorted(files) if f.lower().endswith(extensoes)]

        ano_dict = dados.setdefault(ano, {})
        ano_dict[obra_path] = [{'name': f} for f in arquivos]

    arvore = []
    for ano in sorted(dados.keys()):
        obras = []
        for obra in sorted(dados[ano].keys()):
            obras.append({'name': obra, 'children': dados[ano][obra]})
        arvore.append({'name': ano, 'children': obras})

    return arvore


@bp.route('/api/fotos')
def api_listar_fotos():
    """Lista pastas de fotos, incluindo subpastas vazias."""
    return jsonify(_build_asbuilt_tree(FOTOS_DIR))


@bp.route('/api/fotos/raw/<path:filepath>')
def api_foto_raw(filepath: str):
    filepath = urllib.parse.unquote(filepath)
    try:
        file_path = _safe_join(FOTOS_DIR, *filepath.split('/'))
    except ValueError:
        return jsonify({'error': 'Caminho inválido'}), 400
    if not os.path.isfile(file_path):
        return jsonify({'error': 'Arquivo não encontrado'}), 404
    return send_file(file_path)


@bp.route('/api/fotos/upload', methods=['POST'])
def api_enviar_foto():
    ano = request.form.get('ano', '').strip()
    obra = request.form.get('obra', '').strip()
    arquivo = request.files.get('foto')
    if not ano or not obra or not arquivo:
        return jsonify({'error': 'Dados incompletos'}), 400
    filename = secure_filename(arquivo.filename)
    try:
        destino = _safe_join(FOTOS_DIR, ano, *obra.split('/'))
        os.makedirs(destino, exist_ok=True)
    except ValueError:
        return jsonify({'error': 'Caminho inválido'}), 400
    caminho = os.path.join(destino, filename)
    arquivo.save(caminho)
    return jsonify({'ok': True})    